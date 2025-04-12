import sys
import base64
import hashlib
import csv
import os
import serial
import serial.tools.list_ports

import qdarktheme

from openpyxl import Workbook, load_workbook

from PyQt5.QtWidgets import (QApplication, QMainWindow, QTableWidget, QTableWidgetItem, 
                             QMenu, QMenuBar, QAction, QMessageBox, QFileDialog, 
                             QDialog, QLabel, QLineEdit, QComboBox, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QTextEdit, QStatusBar,QHeaderView,  QGroupBox, QCheckBox,QListWidget,QGraphicsDropShadowEffect, QWidget,QVBoxLayout, QListWidgetItem)
from PyQt5.QtCore import Qt, QSettings
from PyQt5.QtGui import QPalette, QColor

from PyQt5.QtGui import QIcon, QPixmap, QDesktopServices

INI_PATH = os.path.dirname(__file__) + "/ayarlar.ini"
SETTINGS = QSettings(INI_PATH, QSettings.IniFormat)


DEVICE_TYPE_MAP = {
    "00": "Genel amaçlı",
    "01": "Yağ Sayacı",
    "02": "Isı Sayacı",
    "03": "Isı Pay Ölçer (H.C.A)",
    "04": "Kalorimetre",
    "05": "Isı/Soğutma",
    "06": "Sıcak Su",
    "07": "Soğuk Su",
    "08": "Su Sayacı",
    "09": "Gaz Sayacı",
    "0A": "Elektrik Sayacı",
    "0B": "Isı Sayacı",
    "0C": "Soğutma Sayacı",
    "0D": "Kombine Isı/Soğutma",
    "0E": "Basınç Ölçer",
    "0F": "Elektrik",
    "10": "Enerji"
}

DEVICE_TYPE_MAP_REVERSE = {v: k for k, v in DEVICE_TYPE_MAP.items()}

class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Uygulama Ayarları")
        self.setFixedSize(400, 200)

        layout = QVBoxLayout()

        layout.addWidget(QLabel("Excel Yolu:"))

        path_layout = QHBoxLayout()
        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText("Excel yedekleme klasörü")
        self.path_input.setText(SETTINGS.value("excel_export_path", ""))
        browse_button = QPushButton("📁")
        browse_button.clicked.connect(self.select_folder)

        path_layout.addWidget(self.path_input)
        path_layout.addWidget(browse_button)
        layout.addLayout(path_layout)


        self.backup_checkbox = QCheckBox("Otomatik Yedekleme Aktif")
        self.backup_checkbox.setChecked(SETTINGS.value("auto_backup", False, type=bool))
        layout.addWidget(self.backup_checkbox)

        save_btn = QPushButton("Kaydet")
        save_btn.clicked.connect(self.save_settings)
        layout.addWidget(save_btn)

        self.setLayout(layout)

    def save_settings(self):
        SETTINGS.setValue("excel_export_path", self.path_input.text().strip())
        SETTINGS.setValue("auto_backup", self.backup_checkbox.isChecked())
        SETTINGS.sync()
        QMessageBox.information(self, "Ayarlar", "Ayarlar kaydedildi.")
        self.accept()



    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Yedekleme Klasörü Seç")
        if folder:
            self.path_input.setText(folder)


class EnerjiPayMainGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.load_custom_columns()
        self.setWindowTitle("EnerjiPay-RF")
        self.setWindowIcon(QIcon("icon.png"))
        self.resize(1100, 600)

        self.editable = False
        self.modified = False
        self.last_saved_path = None

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Arama yapın...")
        self.search_input.textChanged.connect(self.filter_table)

        self.setup_ui()
        self.create_menus()
        self.setup_status_bar()

        settings = QSettings("EnerjiPay", "EnerjiPay-RF")
        self.dark_mode = settings.value("dark_mode", False, type=bool)
        if self.dark_mode:
            self.apply_dark_theme()
            self.theme_toggle_btn.setText("☀ Açık Tema")
        else:
            self.apply_light_theme()
            self.theme_toggle_btn.setText("🌙 Koyu Tema")

    def apply_custom_light_theme(self):
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor("#fdfdfd"))
        palette.setColor(QPalette.WindowText, Qt.black)
        palette.setColor(QPalette.Base, QColor("#ffffff"))
        palette.setColor(QPalette.AlternateBase, QColor("#f4f4f4"))
        palette.setColor(QPalette.ToolTipBase, Qt.black)
        palette.setColor(QPalette.ToolTipText, Qt.white)
        palette.setColor(QPalette.Text, Qt.black)
        palette.setColor(QPalette.Button, QColor("#e0e0e0"))
        palette.setColor(QPalette.ButtonText, Qt.black)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Highlight, QColor("#3498db"))
        palette.setColor(QPalette.HighlightedText, Qt.white)
        self.setPalette(palette)

        self.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #2471a3;
            }

            QComboBox {
                background-color: white;
                padding: 6px;
                border: 1px solid #ccc;
                border-radius: 5px;
            }

            QComboBox::drop-down {
                border-left: 1px solid #ccc;
            }

            QTableWidget {
                background-color: white;
                alternate-background-color: #f9f9f9;
                gridline-color: #e0e0e0;
                selection-background-color: #3498db;
                selection-color: white;
            }

            QHeaderView::section {
                background-color: #e5e5e5;
                font-weight: bold;
                padding: 1px;
                border: 1px solid #d0d0d0;
            }
        """)
    

    def load_custom_columns(self):
        custom_columns = SETTINGS.value("columns")
        if custom_columns:
            self.columns = custom_columns.split("||")
        else:
            self.columns = [
                "Blok", "Daire No", "Sayaç No", "Sayaç Tipi",
                "Protokol Adı", "Üretici", "Okuma Tarihi",
                "Endeks", "Tüketim", "AES Key"
            ]

    def setup_ui(self):
        layout = QVBoxLayout()
        container = QWidget(self)
        container.setLayout(layout)

        layout.addWidget(self.search_input)

        self.table = QTableWidget()
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels(self.columns)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setDefaultSectionSize(20)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.itemDoubleClicked.connect(self.edit_cell)
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                selection-background-color: #3498db;
                selection-color: white;
            }
            QTableWidget::item:hover {
                background-color: #ecf0f1;
            }
        """)

        # Sütun başlığını sabitle (dikeyde kaymasın)
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.horizontalHeader().setSectionsClickable(True)
        self.table.verticalHeader().setVisible(True)
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)

        layout.addWidget(self.table)

        self.setCentralWidget(container)
        self.load_column_settings()

    def open_app_settings(self):
        dialog = SettingsDialog(self)
        dialog.exec_()

    def filter_table(self, text):
        text = text.strip().lower()
        for row in range(self.table.rowCount()):
            visible = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text in item.text().lower():
                    visible = True
                    break
            self.table.setRowHidden(row, not visible)

    def get_machine_id(self):
        import getpass
        import platform
        import socket
        import uuid

        username = getpass.getuser()              # kullanıcı adı
        hostname = socket.gethostname()           # bilgisayar adı
        system = platform.system()                # işletim sistemi
        release = platform.release()              # sürüm
        node = str(uuid.getnode())                # MAC adresi

        base = f"{username}-{hostname}-{system}-{release}-{node}"
        return "EnerjiPayRF-" + hashlib.sha256(base.encode()).hexdigest()[:16].upper()

    def verify_license(self):
        settings = QSettings("EnerjiPay", "EnerjiPay-RF")
        license_key = settings.value("lisans_kodu", "")
        machine_id = self.get_machine_id()
        expected_key = hashlib.sha256(machine_id.encode()).hexdigest().upper()[:16]
        return license_key == expected_key

    def show_license_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Lisans Anahtarı Girin")
        layout = QVBoxLayout()

        machine_id_edit = QLineEdit(self.get_machine_id())
        machine_id_edit.setReadOnly(True)
        machine_id_edit.setStyleSheet("background-color: #eee;")
        key_input = QLineEdit()
        key_input.setPlaceholderText("Lisans Anahtarı")
        save_btn = QPushButton("Kaydet ve Devam Et")

        layout.addWidget(QLabel("Makine Kimliği:"))
        layout.addWidget(machine_id_edit)
        layout.addWidget(key_input)
        layout.addWidget(save_btn)
        dialog.setLayout(layout)

        if self.verify_license():
            self.dosya_menu.setDisabled(False)
            self.kayitlar_menu.setDisabled(False)
            self.araclar_menu.setDisabled(False)

        def save_key():
            key = key_input.text().strip().upper()
            expected = hashlib.sha256(self.get_machine_id().encode()).hexdigest().upper()[:16]
            if key == expected:
                QSettings("EnerjiPay", "EnerjiPay-RF").setValue("lisans_kodu", key)
                dialog.accept()
            else:
                QMessageBox.critical(dialog, "Hatalı", "Geçersiz lisans anahtarı!")

        save_btn.clicked.connect(save_key)
        dialog.exec_()

    
    def create_menus(self):
        menubar = self.menuBar()

        # --- Dosya Menüsü ---
        self.dosya_menu = menubar.addMenu("Dosya")
        new_action = QAction("Yeni Bina", self)
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(self.clear_table)
        self.dosya_menu.addAction(new_action)

        open_action = QAction("Aç ...", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self.load_syc_file)
        self.dosya_menu.addAction(open_action)

        save_action = QAction("Kaydet ...", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(self.save_syc_file)
        self.dosya_menu.addAction(save_action)

        save_as_action = QAction("Farklı Kaydet ...", self)
        save_as_action.setShortcut("Shift+Ctrl+S")
        save_as_action.triggered.connect(self.save_syc_file_as)
        self.dosya_menu.addAction(save_as_action)

        export_excel_action = QAction("Excel'e Ver", self)
        export_excel_action.triggered.connect(self.export_to_excel)
        self.dosya_menu.addAction(export_excel_action)

        import_excel_action = QAction("Excel'den Al", self)
        import_excel_action.triggered.connect(self.import_from_excel)
        self.dosya_menu.addAction(import_excel_action)

        quit_action = QAction("Kapat", self)
        quit_action.setShortcut("Ctrl+Q")
        quit_action.triggered.connect(self.close)
        self.dosya_menu.addAction(quit_action)

        # --- Kayıtlar Menüsü ---
        self.kayitlar_menu = menubar.addMenu("Kayıtlar")
        edit_mode_action = QAction("Düzenlemeyi Etkinleştir", self)
        edit_mode_action.setShortcut("F2")
        edit_mode_action.triggered.connect(self.toggle_edit_mode)
        self.kayitlar_menu.addAction(edit_mode_action)

        new_record_action = QAction("Yeni Kayıt", self)
        new_record_action.setShortcut("Ctrl+Ins")
        new_record_action.triggered.connect(self.add_empty_row)
        self.kayitlar_menu.addAction(new_record_action)

        delete_record_action = QAction("Kayıt Sil", self)
        delete_record_action.setShortcut("Ctrl+Del")
        delete_record_action.triggered.connect(self.delete_selected_row)
        self.kayitlar_menu.addAction(delete_record_action)

        bulk_add_action = QAction("Toplu Sayaç Ekle", self)
        bulk_add_action.setShortcut("Ctrl+T")
        bulk_add_action.triggered.connect(self.open_bulk_dialog)
        self.kayitlar_menu.addAction(bulk_add_action)

        # --- Araçlar Menüsü ---
        self.araclar_menu = menubar.addMenu("Araçlar")
        settings_action = QAction("Seçenekler", self)
        settings_action.triggered.connect(self.open_settings)
        self.araclar_menu.addAction(settings_action)

        com_settings_action = QAction("Com Ayarları", self)
        com_settings_action.triggered.connect(self.open_com_settings)
        self.araclar_menu.addAction(com_settings_action)

        reset_action = QAction("Varsayılana Dön", self)
        reset_action.triggered.connect(self.reset_to_default)
        self.araclar_menu.addAction(reset_action)

        settings_action2 = QAction("Uygulama Ayarları", self)
        settings_action2.triggered.connect(self.open_app_settings)
        self.araclar_menu.addAction(settings_action2)



        # --- Yardım Menüsü ---
        self.yardim_menu = menubar.addMenu("Yardım")

        about_action = QAction("Hakkında", self)
        about_action.setShortcut("F1")
        about_action.triggered.connect(self.show_about)
        self.yardim_menu.addAction(about_action)

        license_action = QAction("🔐 Lisans Durumu", self)
        license_action.triggered.connect(self.show_license_status)
        self.yardim_menu.addAction(license_action)

        reenter_license_action = QAction("🔐 Lisans Gir", self)
        reenter_license_action.triggered.connect(self.show_license_dialog)
        self.yardim_menu.addAction(reenter_license_action)

        update_action = QAction("Güncelleştirmeleri Denetle", self)
        update_action.triggered.connect(self.check_for_updates)
        self.yardim_menu.addAction(update_action)

        test_telegram_action = QAction("🧪 Test Telgrafı Yolla", self)
        test_telegram_action.triggered.connect(self.send_test_telegram)
        self.yardim_menu.addAction(test_telegram_action)

        logs_action = QAction("📄 Logları Görüntüle", self)
        logs_action.triggered.connect(self.show_logs)
        self.yardim_menu.addAction(logs_action)

        # Lisans yoksa menüleri devre dışı bırak
        if not self.verify_license():
            self.dosya_menu.setDisabled(True)
            self.kayitlar_menu.setDisabled(True)
            self.araclar_menu.setDisabled(True)

    def setup_status_bar(self):
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        

        self.theme_toggle_btn = QPushButton("🌗 Tema")
        self.theme_toggle_btn.setStyleSheet("padding: 4px;")
        self.theme_toggle_btn.clicked.connect(self.toggle_theme)
        self.statusBar.addPermanentWidget(self.theme_toggle_btn)
        self.statusBar.setStyleSheet("""
        QStatusBar {
            background-color: #f4f4f4;
            border-top: 1px solid #d0d0d0;
        }
        QStatusBar::item {
            border: none;
        }
    """)

        # COM Port etiketi
        self.com_label = QLabel("🔌 COM7")
        self.com_label.setContextMenuPolicy(Qt.CustomContextMenu)
        self.com_label.customContextMenuRequested.connect(self.show_com_menu)
        self.statusBar.addPermanentWidget(self.com_label)

        # Bağlantı göstergesi
        self.connect_indicator = QPushButton("✅")
        self.connect_indicator.setStyleSheet("color: green;")
        self.connect_indicator.clicked.connect(self.connect_to_selected_port)
        self.statusBar.addPermanentWidget(self.connect_indicator)

        # Okuma butonu
        self.reading = False
        self.read_button = QPushButton("▶ Okuma Başlat")
        self.read_button.setObjectName("readButton")  # ID veriyoruz
        self.read_button.setStyleSheet("background-color: #28a745; color: white;")
        self.read_button.clicked.connect(self.toggle_reading)
        self.statusBar.addPermanentWidget(self.read_button)
        self.read_button.setStyleSheet("""
        QPushButton#readButton {
            background-color: #27ae60;
            color: white;
            padding: 6px 16px;
            border: none;
            border-radius: 6px;
            font-weight: bold;
            font-size: 14px;
        }
        QPushButton#readButton:hover {
            background-color: #2ecc71;
        }
        QPushButton#readButton:pressed {
            background-color: #1e8449;
        }
    """)
    def show_com_menu(self, pos):
        menu = QMenu()

        ports = [port.device for port in serial.tools.list_ports.comports()]
        if not ports:
            ports = ["COM7"]

        for port in ports:
            action = QAction(port, self)
            action.triggered.connect(lambda checked, p=port: self.set_com_port(p))
            menu.addAction(action)

        menu.exec_(self.com_label.mapToGlobal(pos))
    
    def set_com_port(self, port):
        self.com_label.setText(f"🔌 {port}")


    def apply_dark_theme(self):
        # pyqtdarktheme paketini kullanarak koyu tema stilini yükler
        dark_stylesheet = qdarktheme.load_stylesheet()
        self.setStyleSheet(dark_stylesheet)

    def apply_light_theme(self):
        # Açık temada özel bir stil kullanmıyoruz, dolayısıyla stil dizgesini temizliyoruz
        self.setStyleSheet("")
        self.apply_custom_light_theme()

    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        settings = QSettings("EnerjiPay", "EnerjiPay-RF")
        settings.setValue("dark_mode", self.dark_mode)

        if self.dark_mode:
            self.apply_dark_theme()
            self.theme_toggle_btn.setText("☀ Açık Tema")
        else:
            self.apply_light_theme()
            self.theme_toggle_btn.setText("🌙 Koyu Tema")
    def toggle_reading(self):
        if not getattr(self, "serial_connected", False):
            QMessageBox.warning(self, "Uyarı", "COM bağlantısı yapılmadan okuma başlatılamaz.")
            return

        self.reading = not self.reading
        if self.reading:
            self.read_button.setText("■ Okumayı Durdur")
            self.read_button.setStyleSheet("background-color: #dc3545; color: white; font-weight: bold;")
            print("Okuma başlatıldı")
        else:
            self.read_button.setText("▶ Okuma Başlat")
            self.read_button.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
            print("Okuma durduruldu")


            self.reading = not self.reading
        if self.reading:
            self.read_button.setText("■ Okumayı Durdur")
            self.read_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    font-weight: bold;
                    padding: 8px 16px;
                    border-radius: 8px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            print("Okuma başlatıldı")
        else:
            self.read_button.setText("▶ Okuma Başlat")
            self.read_button.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    font-weight: bold;
                    padding: 8px 16px;
                    border-radius: 8px;
                }
                QPushButton:hover {
                    background-color: #218838;
                }
            """)
            print("Okuma durduruldu")

    

    def edit_cell(self, item):
        if not self.editable:
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Hücreyi Düzenle")
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f4f4f4;
                border-radius: 10px;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 8px 15px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Hücre etiketi
        column_name = self.columns[item.column()]
        label = QLabel(f"<b>{column_name}</b> sütununu düzenliyorsunuz:")
        layout.addWidget(label)
        
        edit_line = QLineEdit(item.text())
        edit_line.selectAll()  # Tüm metni seç
        layout.addWidget(edit_line)
        
        save_button = QPushButton("Kaydet")
        save_button.clicked.connect(lambda: self.save_cell_edit(item, edit_line.text(), dialog))
        layout.addWidget(save_button)
        
        dialog.setLayout(layout)
        dialog.exec_()

    def save_cell_edit(self, item, new_text, dialog):
        item.setText(new_text)
        dialog.accept()

    def toggle_edit_mode(self):
        self.editable = not self.editable
        state = "aktif" if self.editable else "pasif"
        QMessageBox.information(self, "Düzenleme", f"Düzenleme modu {state} hale getirildi.")

    def add_empty_row(self):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        self.modified = True

    def delete_selected_row(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.table.removeRow(current_row)

    def open_bulk_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Toplu Sayaç Ekle")
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f4f4f4;
            }
            QLabel {
                font-weight: bold;
            }
            QLineEdit, QComboBox, QTextEdit {
                padding: 6px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
            }
            QPushButton {
                background-color: #2ecc71;
                color: white;
                padding: 8px 15px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        dialog.setFixedSize(500, 350)

        main_layout = QVBoxLayout()
        
        card = QGroupBox("Toplu Sayaç Bilgileri")
        card_layout = QVBoxLayout()
        card.setLayout(card_layout)
        card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #dcdcdc;
                border-radius: 8px;
                padding: 12px;
                font-weight: bold;
            }
""")

        layout = QVBoxLayout()

        daire_layout = QHBoxLayout()
        daire_label = QLabel("Başlangıç daire No:")
        daire_entry = QLineEdit()
        daire_layout.addWidget(daire_label)
        daire_layout.addWidget(daire_entry)
        layout.addLayout(daire_layout)

        tip_layout = QHBoxLayout()
        tip_label = QLabel("Sayaç Tipi:")
        tip_combo = QComboBox()
        tip_combo.addItems(["Kalorimetre", "Sıcak Su", "Soğuk Su"])
        tip_layout.addWidget(tip_label)
        tip_layout.addWidget(tip_combo)
        layout.addLayout(tip_layout)

        protokol_layout = QHBoxLayout()
        protokol_label = QLabel("Protokol:")
        protokol_combo = QComboBox()
        protokol_combo.addItems(["Standart M-Bus"])
        protokol_layout.addWidget(protokol_label)
        protokol_layout.addWidget(protokol_combo)
        layout.addLayout(protokol_layout)

        sayaç_label = QLabel("Sayaçlar:")
        sayaç_text = QTextEdit()
        sayaç_text.setPlaceholderText("Her satıra bir sayaç numarası yazın")
        layout.addWidget(sayaç_label)
        layout.addWidget(sayaç_text)

        aciklama_label = QLabel("Saptaki bölüme her satıra bir adet olacak şekilde sayaç numaralarını yazın")
        layout.addWidget(aciklama_label)

        ekle_button = QPushButton("Ekle")
        ekle_button.clicked.connect(lambda: self.bulk_add_counters(
            daire_entry.text(), 
            tip_combo.currentText(), 
            protokol_combo.currentText(), 
            sayaç_text.toPlainText().split('\n'), 
            dialog
        ))
        layout.addWidget(ekle_button)

        dialog.setLayout(layout)
        dialog.exec_()

    def bulk_add_counters(self, start_daire, tip, protokol, sayaçlar, dialog):
        try:
            start = int(start_daire)
        except ValueError:
            QMessageBox.critical(self, "Hata", "Başlangıç daire numarası geçerli değil.")
            return

        for i, sayaç_no in enumerate(sayaçlar):
            if not sayaç_no.strip():
                continue
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)
            
            row_data = ["A", str(start + i), sayaç_no, tip, protokol] + ["" for _ in range(len(self.columns)-5)]
            
            for j, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                self.table.setItem(row_position, j, item)

        dialog.accept()

    def clear_table(self):
        if self.table.rowCount() > 0:
            confirm = QMessageBox.question(self, "Tabloyu Temizle", "Tabloyu tamamen temizlemek istediğinize emin misiniz?", QMessageBox.Yes | QMessageBox.No)
            if confirm != QMessageBox.Yes:
                return
        self.table.setRowCount(0)


    def save_syc_file(self):
        if self.last_saved_path:
            self._write_syc(self.last_saved_path)
        else:
            self.save_syc_file_as()

    def save_syc_file_as(self):
        path, _ = QFileDialog.getSaveFileName(self, "Sayaç Dosyası Kaydet", "", "Sayaç Dosyası (*.syc)")
        if path:
            self.last_saved_path = path
            self._write_syc(path)

            # Otomatik yedekleme
            backup_path = SETTINGS.value("excel_export_path", "")
            auto_backup = SETTINGS.value("auto_backup", False, type=bool)
            if auto_backup and backup_path:
                import shutil, datetime
                now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_file = os.path.join(backup_path, f"yedek_{now}.syc")
                try:
                    shutil.copy(path, backup_file)
                except Exception as e:
                    QMessageBox.warning(self, "Yedekleme Hatası", f"Yedekleme başarısız oldu:\n{e}")


    def _write_syc(self, path):
        with open(path, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter='|')
            writer.writerow(self.columns)
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    value = item.text() if item else ""
                    if self.columns[col] == "Sayaç Tipi":
                        for code, label in DEVICE_TYPE_MAP.items():
                            if label == value:
                                value = str(int(code, 16))
                                break
                    row_data.append(value)
                writer.writerow(row_data)

        # 📁 Otomatik Yedekleme
        backup_path = SETTINGS.value("excel_export_path", "")
        auto_backup = SETTINGS.value("auto_backup", False, type=bool)
        if auto_backup and backup_path:
            import shutil
            from datetime import datetime

            name, ext = os.path.splitext(os.path.basename(path))
            tarih = datetime.now().strftime("%d-%m-%Y")
            backup_file = os.path.join(backup_path, f"{name}-{tarih}{ext}")

            try:
                shutil.copy(path, backup_file)
            except Exception as e:
                QMessageBox.warning(self, "Yedekleme Hatası", f"Yedekleme başarısız oldu:\n{e}")

        QMessageBox.information(self, "Kaydedildi", f"Veriler kaydedildi: {os.path.basename(path)}")
        self.modified = False
    def load_syc_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Sayaç Dosyası Aç", "", "Sayaç Dosyası (*.syc)")
        if path:
            self.load_syc_file_from_path(path)

    def load_syc_file_from_path(self, path):
        if os.path.exists(path):
            self.clear_table()
            with open(path, mode='r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter='|')
                headers = next(reader)
                for i, row in enumerate(reader):
                    clean_row = ["" if cell == "None" else cell for cell in row]
                    row_position = self.table.rowCount()
                    self.table.insertRow(row_position)
                    for j, value in enumerate(clean_row):
                        if j < len(self.columns) and self.columns[j] == "Sayaç Tipi":
                            hex_code = value.zfill(2).upper()
                            value = DEVICE_TYPE_MAP.get(hex_code, value)
                        item = QTableWidgetItem(str(value))
                        self.table.setItem(row_position, j, item)
            self.last_saved_path = path
            SETTINGS.setValue("last_opened_file", path)
            SETTINGS.sync()
            self.modified = False

    def export_to_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Excel Dosyası Kaydet", "", "Excel Dosyası (*.xlsx)")
        
        if path:
            wb = Workbook()
            ws = wb.active
            ws.append(self.columns)
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            wb.save(path)
            QMessageBox.information(self, "Aktarıldı", f"Veriler Excel'e aktarıldı: {os.path.basename(path)}")

            # Otomatik yedekleme
            backup_path = SETTINGS.value("excel_export_path", "")
            auto_backup = SETTINGS.value("auto_backup", False, type=bool)

            if auto_backup and backup_path:
                import shutil
                from datetime import datetime

                base = os.path.basename(path)  # örnek: sycliste.xlsx
                name, ext = os.path.splitext(base)  # name=sycliste, ext=.xlsx
                tarih = datetime.now().strftime("%d-%m-%Y")
                backup_file = os.path.join(backup_path, f"{name}-{tarih}{ext}")

                try:
                    shutil.copy(path, backup_file)
                except Exception as e:
                    QMessageBox.warning(self, "Yedekleme Hatası", f"Yedekleme başarısız oldu:\n{e}")



    def import_from_excel(self):
        if self.table.rowCount() > 0:
            confirm = QMessageBox.question(self, "Excel'den Al", "Mevcut veriler silinecek. Devam edilsin mi?", QMessageBox.Yes | QMessageBox.No)
            if confirm != QMessageBox.Yes:
                return
        path, _ = QFileDialog.getOpenFileName(self, "Excel Dosyası Aç", "", "Excel Dosyası (*.xlsx)")
        if path:
            self.clear_table()
            wb = load_workbook(path)
            ws = wb.active

            excel_headers = []
            data_rows = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    excel_headers = [str(cell).strip() if cell else "" for cell in row]
                else:
                    data_rows.append(["" if cell is None else str(cell) for cell in row])

            header_mapping = {}
            for table_index, column_name in enumerate(self.columns):
                for excel_index, excel_header in enumerate(excel_headers):
                    if column_name.lower() == excel_header.lower():
                        header_mapping[table_index] = excel_index
                        break

            for row_data in data_rows:
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                for col_index, excel_index in header_mapping.items():
                    if excel_index < len(row_data):
                        value = row_data[excel_index]
                        if self.columns[col_index] == "Sayaç Tipi":
                            hex_code = value.zfill(2).upper()
                            value = DEVICE_TYPE_MAP.get(hex_code, value)
                        item = QTableWidgetItem(str(value))
                        self.table.setItem(row_position, col_index, item)

    def connect_to_selected_port(self):
        port = self.com_label.text().replace("🔌", "").strip()
        try:
            self.serial_port = serial.Serial(port, baudrate=115200, timeout=1)
            self.serial_connected = True  # Bağlantı başarılı
            QMessageBox.information(self, "Bağlantı", f"{port} portuna başarıyla bağlanıldı.")
        except Exception as e:
            self.serial_connected = False  # Bağlantı başarısız
            QMessageBox.critical(self, "Hata", f"Bağlantı başarısız:\n{str(e)}")



    def open_com_settings(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("COM Ayarları")
        dialog.setFixedSize(350, 220)
        dialog.setStyleSheet("QDialog { background-color: #f9f9f9; }")

        main_layout = QVBoxLayout()

        card = QGroupBox("Port Ayarları")
        card.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 8px;
                padding: 12px;
                font-weight: bold;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15)
        shadow.setXOffset(0)
        shadow.setYOffset(3)
        shadow.setColor(Qt.gray)
        card.setGraphicsEffect(shadow)

        card_layout = QVBoxLayout()

        available_ports = [port.device for port in serial.tools.list_ports.comports()]
        if not available_ports:
            available_ports = ["COM7"]

        port_layout = QHBoxLayout()
        port_label = QLabel("Port:")
        port_combo = QComboBox()
        port_combo.addItems(available_ports)
        port_combo.setCurrentText("COM7" if "COM7" in available_ports else available_ports[0])
        port_layout.addWidget(port_label)
        port_layout.addWidget(port_combo)
        card_layout.addLayout(port_layout)

        baud_layout = QHBoxLayout()
        baud_label = QLabel("Baudrate:")
        baud_combo = QComboBox()
        baud_combo.addItems(["9600", "19200", "38400", "57600", "115200"])
        baud_combo.setCurrentText("115200")
        baud_layout.addWidget(baud_label)
        baud_layout.addWidget(baud_combo)
        card_layout.addLayout(baud_layout)

        connect_button = QPushButton("🔌 Bağlan")
        connect_button.clicked.connect(lambda: self.test_com_connection(
            port_combo.currentText(), int(baud_combo.currentText())
        ))
        card_layout.addWidget(connect_button)

        card.setLayout(card_layout)
        main_layout.addWidget(card)
        dialog.setLayout(main_layout)
        dialog.exec_()

    def test_com_connection(self, port, baudrate):
        try:
            ser = serial.Serial(port, baudrate=baudrate, timeout=1)
            QMessageBox.information(self, "Bağlantı", f"{port} portuna başarıyla bağlanıldı.")
            self.com_label.setText(f"🔌 {port}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Bağlantı başarısız:\n{str(e)}")

    def open_settings(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Sütun Ayarları")
        dialog.setFixedSize(400, 400)
        layout = QVBoxLayout()

        # Sütun listesi
        self.column_list = QListWidget()
        self.column_list.setDragDropMode(QListWidget.InternalMove)  # Sürükle-bırak sıralama
        self.column_list.setSelectionMode(QListWidget.SingleSelection)
        self.column_list.addItems(self.columns)
        layout.addWidget(QLabel("Sıralamak için sürükleyin, silmek için seçip 'Sil' deyin:"))
        layout.addWidget(self.column_list)

        # Sütun ekleme alanı
        add_layout = QHBoxLayout()
        add_input = QLineEdit()
        add_input.setPlaceholderText("Yeni sütun adı")
        add_button = QPushButton("➕ Ekle")
        add_button.clicked.connect(lambda: self.add_new_column(add_input.text()))
        add_layout.addWidget(add_input)
        add_layout.addWidget(add_button)
        layout.addLayout(add_layout)

        # Sil butonu
        delete_button = QPushButton("🗑 Seçili Sütunu Sil")
        delete_button.clicked.connect(self.delete_selected_column)
        layout.addWidget(delete_button)

        # Kaydet butonu
        save_button = QPushButton("💾 Uygula")
        save_button.clicked.connect(lambda: self.apply_column_settings(dialog))
        layout.addWidget(save_button)

        dialog.setLayout(layout)
        dialog.exec_()

    def delete_selected_column(self):
        selected = self.column_list.currentRow()
        if selected >= 0:
            self.column_list.takeItem(selected)

    def add_new_column(self, column_name):
        if column_name.strip():
            self.column_list.addItem(column_name.strip())

    def apply_column_settings(self, dialog):
        new_columns = [self.column_list.item(i).text() for i in range(self.column_list.count())]
        self.columns = new_columns
        self.save_custom_columns()
        self.save_column_settings()
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels(self.columns)
        dialog.accept()


    def move_column_item(self, direction):
        row = self.column_list.currentRow()
        if row < 0 or row + direction < 0 or row + direction >= self.column_list.count():
            return
        item = self.column_list.takeItem(row)
        self.column_list.insertItem(row + direction, item)
        self.column_list.setCurrentRow(row + direction)

    def save_column_settings(self, dialog):
        new_columns = []
        for i in range(self.column_list.count()):
            item = self.column_list.item(i)
            if item.checkState() == Qt.Checked:
                new_columns.append(item.text())
        if new_columns:
            self.columns = new_columns
            self.setup_ui()
        dialog.accept()

    def show_license_status(self):
        if self.verify_license():
            QMessageBox.information(self, "Lisans Durumu", "Bu sistem lisanslıdır. ✔")
        else:
            QMessageBox.warning(self, "Lisans Durumu", "Lisans geçersiz. Lütfen lisans anahtarınızı girin.")

    def check_for_updates(self):
        QMessageBox.information(self, "Güncelleme", "Şu anda en güncel sürümü kullanıyorsun.")

    def send_test_telegram(self):
        QMessageBox.information(self, "Test Telgrafı", "Test telgrafı gönderme işlevi henüz tamamlanmadı.")

    def show_logs(self):
        log_path = "wmbus_logs.txt"
        try:
            with open(log_path, "r", encoding="utf-8") as f:
                content = f.read()
            log_dialog = QDialog(self)
            log_dialog.setWindowTitle("Loglar")
            log_dialog.resize(600, 400)
            layout = QVBoxLayout()
            log_text = QTextEdit()
            log_text.setPlainText(content)
            log_text.setReadOnly(True)
            layout.addWidget(log_text)
            log_dialog.setLayout(layout)
            log_dialog.exec_()
        except FileNotFoundError:
            QMessageBox.information(self, "Log", "Log dosyası bulunamadı.")

    def show_about(self):
        about_text = """
        EnerjiPay-RF v1.0

        Bu uygulama, enerji sayaçlarının okuma ve yönetimi için tasarlanmış bir araçtır.
        wM-Bus protokolünü kullanarak kablosuz sayaçlardan veri toplayabilir ve verilerinizi analiz edebilirsiniz.

        Geliştirici:
        - Nortek Aypa Otomasyon
            0505 059 88 58
        Lisans: GPL-3.0
        """
        about_dialog = QDialog(self)
        about_dialog.setWindowTitle("Hakkında")
        about_dialog.setMinimumSize(300, 150)
        layout = QVBoxLayout()
        logo_label = QLabel()
        pixmap = QPixmap("logo2.png")
        scaled_pixmap = pixmap.scaled(150, 50, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        logo_label.setPixmap(scaled_pixmap)
        logo_label.setAlignment(Qt.AlignCenter)
        logo_label.setCursor(Qt.PointingHandCursor)
        logo_label.mousePressEvent = lambda event: self.open_website()
        layout.addWidget(logo_label)
        text_label = QLabel(about_text)
        text_label.setWordWrap(True)
        text_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(text_label)
        close_button = QPushButton("Kapat")
        close_button.clicked.connect(about_dialog.close)
        layout.addWidget(close_button)
        about_dialog.setLayout(layout)
        about_dialog.exec_()

    def open_website(self):
        import webbrowser
        webbrowser.open('https://www.enerjipay.com')

    def save_column_settings(self):
        settings = QSettings("EnerjiPay", "EnerjiPay-RF")
        visible_columns = []
        column_order = []

        for i in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(i).text()
            if not self.table.isColumnHidden(i):
                visible_columns.append(header)
            column_order.append(header)

        settings.setValue("visible_columns", visible_columns)
        settings.setValue("column_order", column_order)

    def load_column_settings(self):
        settings = QSettings("EnerjiPay", "EnerjiPay-RF")
        saved_order = settings.value("column_order", self.columns)
        visible_columns = settings.value("visible_columns", self.columns)

        # Sıralamayı uygula
        current_headers = {self.table.horizontalHeaderItem(i).text(): i for i in range(self.table.columnCount())}
        new_order = [current_headers[col] for col in saved_order if col in current_headers]

        for logical_index, visual_index in enumerate(new_order):
            self.table.horizontalHeader().moveSection(self.table.horizontalHeader().visualIndex(visual_index), logical_index)

        # Görünürlüğü uygula
        for i in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(i).text()
            self.table.setColumnHidden(i, header not in visible_columns)

    def closeEvent(self, event):
        self.save_column_settings()
        event.accept()

    def reset_to_default(self):
        settings = QSettings("EnerjiPay", "EnerjiPay-RF")
        settings.remove("column_order")
        settings.remove("visible_columns")
        settings.remove("columns")  # ← bunu ekle!

        # columns listesini varsayılana döndür
        self.columns = [
            "Blok", "Daire No", "Sayaç No", "Sayaç Tipi",
            "Protokol Adı", "Üretici", "Okuma Tarihi",
            "Endeks", "Tüketim", "AES Key"
        ]
        self.save_custom_columns()

        QMessageBox.information(self, "Sıfırlandı", "Sütun ayarları varsayılana döndürüldü. Değişiklikler yeniden başlattığınızda uygulanacaktır.")

    
    def save_custom_columns(self):
        SETTINGS.setValue("columns", "||".join(self.columns))
        SETTINGS.sync()



def main():
    app = QApplication(sys.argv)
    settings = QSettings("EnerjiPay", "EnerjiPay-RF")
    dark_mode = settings.value("dark_mode", False, type=bool)
    main_window = EnerjiPayMainGUI()


    last_file = SETTINGS.value("last_opened_file", "")
    if last_file and os.path.exists(last_file):
        main_window.load_syc_file_from_path(last_file)  # <-- sadece path veriyoruz
    # NOT: Burada load_syc_file() değil, load_syc_file_from_path(path) kullanıyoruz

    if dark_mode:
        main_window.apply_dark_theme()
        main_window.theme_toggle_btn.setText("☀ Açık Tema")
    else:
        main_window.apply_light_theme()
        main_window.theme_toggle_btn.setText("🌙 Koyu Tema")

    main_window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
